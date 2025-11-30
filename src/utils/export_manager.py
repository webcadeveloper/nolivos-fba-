"""
Export Manager - Exporta datos a CSV, Excel, PDF
"""
import csv
import io
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)


class ExportManager:
    """Maneja exportación de datos a diferentes formatos"""

    @staticmethod
    def export_to_csv(data, filename=None):
        """
        Exporta lista de dicts a CSV

        Args:
            data: Lista de diccionarios
            filename: Nombre del archivo (opcional)

        Returns:
            str: Contenido CSV
        """
        if not data:
            return ""

        output = io.StringIO()
        fieldnames = data[0].keys()

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

        csv_content = output.getvalue()
        output.close()

        if filename:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)
            logging.info(f"✅ CSV exportado a {filename}")

        return csv_content

    @staticmethod
    def export_opportunities_to_csv(opportunities, filename=None):
        """Exporta oportunidades a CSV con formato optimizado"""

        if not opportunities:
            return ""

        # Preparar datos con columnas específicas
        export_data = []

        for opp in opportunities:
            export_data.append({
                'ASIN': opp.get('asin'),
                'Producto': opp.get('product_name', '')[:100],
                'Categoría': opp.get('category'),
                'Precio Amazon': f"${opp.get('amazon_price', 0):.2f}",
                'Precio Proveedor': f"${opp.get('supplier_price', 0):.2f}",
                'Proveedor': opp.get('supplier_name'),
                'Costo Total': f"${opp.get('total_cost', 0):.2f}",
                'Fees Amazon': f"${opp.get('amazon_fees', 0):.2f}",
                'Ganancia Neta': f"${opp.get('net_profit', 0):.2f}",
                'ROI %': f"{opp.get('roi_percent', 0):.1f}%",
                'Margen %': f"{opp.get('margin_percent', 0):.1f}%",
                'Nivel': opp.get('competitiveness_level'),
                'BSR': opp.get('bsr', 'N/A'),
                'Ventas Estimadas/Mes': opp.get('estimated_monthly_sales', 0),
                'Fecha Escaneo': opp.get('scan_date'),
                'URL': f"https://www.amazon.com/dp/{opp.get('asin')}"
            })

        return ExportManager.export_to_csv(export_data, filename)

    @staticmethod
    def export_alerts_to_csv(alerts, filename=None):
        """Exporta alertas a CSV"""

        if not alerts:
            return ""

        export_data = []

        for alert in alerts:
            export_data.append({
                'Tipo': alert.get('alert_type'),
                'Severidad': alert.get('severity').upper(),
                'Mensaje': alert.get('message'),
                'Producto': alert.get('product_name', '')[:100],
                'ASIN': alert.get('asin'),
                'Fecha': alert.get('created_at'),
                'Leída': 'Sí' if alert.get('is_read') else 'No'
            })

        return ExportManager.export_to_csv(export_data, filename)

    @staticmethod
    def export_trends_to_csv(trends, filename=None):
        """Exporta tendencias a CSV"""

        if not trends:
            return ""

        export_data = []

        for trend in trends:
            export_data.append({
                'ASIN': trend.get('asin'),
                'Producto': trend.get('product_name', '')[:100],
                'Categoría': trend.get('category'),
                'BSR Actual': trend.get('current_bsr'),
                'Cambio BSR 30d': trend.get('bsr_change_30d'),
                'Tendencia BSR': trend.get('bsr_trend'),
                'Tendencia Demanda': trend.get('demand_trend'),
                'Precio Actual': f"${trend.get('current_price', 0):.2f}",
                'Cambio Precio 30d': f"${trend.get('price_change_30d', 0):.2f}",
                'Sellers Actuales': trend.get('current_sellers'),
                'Cambio Sellers 30d': trend.get('seller_change_30d'),
                'Opportunity Score': trend.get('opportunity_score'),
                'Recomendación IA': trend.get('ai_recommendation', 'N/A')
            })

        return ExportManager.export_to_csv(export_data, filename)

    @staticmethod
    def generate_excel_from_csv(csv_content, filename):
        """
        Convierte CSV a Excel (requiere openpyxl)

        Args:
            csv_content: Contenido CSV
            filename: Nombre archivo Excel (.xlsx)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import csv
            import io

            # Parse CSV
            csv_reader = csv.reader(io.StringIO(csv_content))
            rows = list(csv_reader)

            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Datos"

            # Escribir datos
            for row_idx, row in enumerate(rows, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Estilo para header
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar
            wb.save(filename)
            logging.info(f"✅ Excel exportado a {filename}")

            return True

        except ImportError:
            logging.warning("⚠️ openpyxl no instalado. Usa: pip install openpyxl")
            return False
        except Exception as e:
            logging.error(f"❌ Error generando Excel: {e}")
            return False

    @staticmethod
    def export_opportunities_to_excel(opportunities, filename):
        """
        Exporta oportunidades a Excel con formato profesional

        Args:
            opportunities: Lista de oportunidades
            filename: Nombre archivo Excel (.xlsx)

        Returns:
            bool: True si éxito
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.utils import get_column_letter

            if not opportunities:
                return False

            # Crear workbook
            wb = openpyxl.Workbook()

            # ==================== HOJA 1: SUMMARY ====================
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Calcular estadísticas
            total_opps = len(opportunities)
            avg_roi = sum(o.get('roi_percent', 0) for o in opportunities) / total_opps if total_opps > 0 else 0
            avg_profit = sum(o.get('net_profit', 0) for o in opportunities) / total_opps if total_opps > 0 else 0
            best_opp = max(opportunities, key=lambda x: x.get('roi_percent', 0)) if opportunities else None
            total_potential_profit = sum(o.get('net_profit', 0) for o in opportunities)

            # Título
            ws_summary['A1'] = '📊 REPORTE DE OPORTUNIDADES FBA'
            ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.merge_cells('A1:D1')
            ws_summary.row_dimensions[1].height = 30

            # Fecha
            ws_summary['A2'] = f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws_summary['A2'].font = Font(italic=True)
            ws_summary.merge_cells('A2:D2')

            # Estadísticas
            stats_row = 4
            stats = [
                ('Total Oportunidades:', total_opps),
                ('ROI Promedio:', f'{avg_roi:.1f}%'),
                ('Ganancia Promedia:', f'${avg_profit:.2f}'),
                ('Ganancia Potencial Total:', f'${total_potential_profit:.2f}')
            ]

            for idx, (label, value) in enumerate(stats, start=stats_row):
                ws_summary[f'A{idx}'] = label
                ws_summary[f'A{idx}'].font = Font(bold=True)
                ws_summary[f'B{idx}'] = value
                ws_summary[f'B{idx}'].font = Font(size=12)

            # Mejor oportunidad
            if best_opp:
                best_row = stats_row + len(stats) + 1
                ws_summary[f'A{best_row}'] = '🏆 MEJOR OPORTUNIDAD:'
                ws_summary[f'A{best_row}'].font = Font(bold=True, color="FFFFFF")
                ws_summary[f'A{best_row}'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                ws_summary.merge_cells(f'A{best_row}:D{best_row}')

                best_row += 1
                ws_summary[f'A{best_row}'] = 'ASIN:'
                ws_summary[f'B{best_row}'] = best_opp.get('asin')
                best_row += 1
                ws_summary[f'A{best_row}'] = 'Producto:'
                ws_summary[f'B{best_row}'] = best_opp.get('product_name', '')[:50]
                best_row += 1
                ws_summary[f'A{best_row}'] = 'ROI:'
                ws_summary[f'B{best_row}'] = f"{best_opp.get('roi_percent', 0):.1f}%"
                ws_summary[f'B{best_row}'].font = Font(bold=True, color="4CAF50", size=14)
                best_row += 1
                ws_summary[f'A{best_row}'] = 'Ganancia:'
                ws_summary[f'B{best_row}'] = f"${best_opp.get('net_profit', 0):.2f}"
                ws_summary[f'B{best_row}'].font = Font(bold=True, color="4CAF50", size=14)

            # Ajustar anchos
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 15
            ws_summary.column_dimensions['D'].width = 15

            # ==================== HOJA 2: OPORTUNIDADES ====================
            ws_data = wb.create_sheet("Oportunidades")

            # Headers
            headers = [
                'ASIN', 'Producto', 'Categoría', 'Precio Amazon', 'Precio Proveedor',
                'Proveedor', 'Costo Total', 'Fees Amazon', 'Ganancia Neta',
                'ROI %', 'Margen %', 'Nivel', 'BSR', 'Ventas Est/Mes', 'URL'
            ]

            # Escribir headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    bottom=Side(style='thick', color="000000")
                )

            ws_data.row_dimensions[1].height = 30

            # Escribir datos
            for row_idx, opp in enumerate(opportunities, 2):
                ws_data.cell(row=row_idx, column=1, value=opp.get('asin'))
                ws_data.cell(row=row_idx, column=2, value=opp.get('product_name', '')[:100])
                ws_data.cell(row=row_idx, column=3, value=opp.get('category'))
                ws_data.cell(row=row_idx, column=4, value=opp.get('amazon_price', 0))
                ws_data.cell(row=row_idx, column=5, value=opp.get('supplier_price', 0))
                ws_data.cell(row=row_idx, column=6, value=opp.get('supplier_name'))
                ws_data.cell(row=row_idx, column=7, value=opp.get('total_cost', 0))
                ws_data.cell(row=row_idx, column=8, value=opp.get('amazon_fees', 0))
                ws_data.cell(row=row_idx, column=9, value=opp.get('net_profit', 0))
                ws_data.cell(row=row_idx, column=10, value=opp.get('roi_percent', 0))
                ws_data.cell(row=row_idx, column=11, value=opp.get('margin_percent', 0))
                ws_data.cell(row=row_idx, column=12, value=opp.get('competitiveness_level', ''))
                ws_data.cell(row=row_idx, column=13, value=opp.get('bsr', ''))
                ws_data.cell(row=row_idx, column=14, value=opp.get('estimated_monthly_sales', 0))
                ws_data.cell(row=row_idx, column=15, value=f"https://www.amazon.com/dp/{opp.get('asin')}")

                # Formato de números
                ws_data.cell(row=row_idx, column=4).number_format = '$#,##0.00'
                ws_data.cell(row=row_idx, column=5).number_format = '$#,##0.00'
                ws_data.cell(row=row_idx, column=7).number_format = '$#,##0.00'
                ws_data.cell(row=row_idx, column=8).number_format = '$#,##0.00'
                ws_data.cell(row=row_idx, column=9).number_format = '$#,##0.00'
                ws_data.cell(row=row_idx, column=10).number_format = '0.0"%"'
                ws_data.cell(row=row_idx, column=11).number_format = '0.0"%"'

            # Conditional Formatting para ROI
            roi_col = get_column_letter(10)  # Columna J (ROI %)

            # Verde: ROI > 50%
            ws_data.conditional_formatting.add(
                f'{roi_col}2:{roi_col}{len(opportunities)+1}',
                CellIsRule(operator='greaterThan', formula=['50'],
                          fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
                          font=Font(bold=True, color="2E7D32"))
            )

            # Amarillo: ROI 20-50%
            ws_data.conditional_formatting.add(
                f'{roi_col}2:{roi_col}{len(opportunities)+1}',
                CellIsRule(operator='between', formula=['20', '50'],
                          fill=PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
                          font=Font(bold=True, color="F57F17"))
            )

            # Rojo: ROI < 20%
            ws_data.conditional_formatting.add(
                f'{roi_col}2:{roi_col}{len(opportunities)+1}',
                CellIsRule(operator='lessThan', formula=['20'],
                          fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
                          font=Font(bold=True, color="C62828"))
            )

            # Auto-ajustar anchos
            column_widths = {
                'A': 12, 'B': 40, 'C': 15, 'D': 13, 'E': 15,
                'F': 15, 'G': 12, 'H': 12, 'I': 13, 'J': 10,
                'K': 10, 'L': 18, 'M': 12, 'N': 15, 'O': 35
            }

            for col, width in column_widths.items():
                ws_data.column_dimensions[col].width = width

            # ==================== GRÁFICO 1: TOP 10 ROI ====================
            if len(opportunities) >= 3:
                ws_charts = wb.create_sheet("Gráficos")

                # Preparar datos para gráfico (top 10 por ROI)
                top_10 = sorted(opportunities, key=lambda x: x.get('roi_percent', 0), reverse=True)[:10]

                # Escribir datos para gráfico
                ws_charts['A1'] = 'Producto'
                ws_charts['B1'] = 'ROI %'

                for idx, opp in enumerate(top_10, 2):
                    ws_charts[f'A{idx}'] = opp.get('product_name', '')[:30]
                    ws_charts[f'B{idx}'] = opp.get('roi_percent', 0)

                # Crear gráfico de barras
                chart = BarChart()
                chart.title = "Top 10 Oportunidades por ROI"
                chart.x_axis.title = "Producto"
                chart.y_axis.title = "ROI %"
                chart.style = 10
                chart.height = 12
                chart.width = 25

                data = Reference(ws_charts, min_col=2, min_row=1, max_row=len(top_10)+1)
                cats = Reference(ws_charts, min_col=1, min_row=2, max_row=len(top_10)+1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                ws_charts.add_chart(chart, "D2")

                # ==================== GRÁFICO 2: DISTRIBUCIÓN POR CATEGORÍA ====================
                # Contar por categoría
                from collections import Counter
                category_counts = Counter(o.get('category', 'Unknown') for o in opportunities)

                # Escribir datos para pie chart
                start_row = len(top_10) + 5
                ws_charts[f'A{start_row}'] = 'Categoría'
                ws_charts[f'B{start_row}'] = 'Cantidad'

                for idx, (cat, count) in enumerate(category_counts.items(), start_row+1):
                    ws_charts[f'A{idx}'] = cat
                    ws_charts[f'B{idx}'] = count

                # Crear pie chart
                pie = PieChart()
                pie.title = "Distribución por Categoría"
                pie.style = 10
                pie.height = 12
                pie.width = 25

                labels = Reference(ws_charts, min_col=1, min_row=start_row+1, max_row=start_row+len(category_counts))
                data = Reference(ws_charts, min_col=2, min_row=start_row, max_row=start_row+len(category_counts))

                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)

                ws_charts.add_chart(pie, f"D{start_row+2}")

            # Guardar
            wb.save(filename)
            logging.info(f"✅ Excel profesional exportado a {filename}")

            return True

        except ImportError as e:
            logging.warning(f"⚠️ Librería faltante: {e}. Usa: pip install openpyxl")
            return False
        except Exception as e:
            logging.error(f"❌ Error generando Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    @staticmethod
    def create_download_response(csv_content, filename):
        """
        Crea respuesta Flask para descarga

        Args:
            csv_content: Contenido CSV
            filename: Nombre del archivo

        Returns:
            Flask response
        """
        from flask import make_response

        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    @staticmethod
    def create_excel_download_response(filepath, filename):
        """
        Crea respuesta Flask para descarga de Excel

        Args:
            filepath: Path al archivo Excel
            filename: Nombre para descarga

        Returns:
            Flask response
        """
        from flask import send_file

        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
